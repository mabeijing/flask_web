from openpyxl import load_workbook
wb = load_workbook('excelData.xlsx')
ws = wb.active
# 方法一
# c = ws['D3'].value
# print(c)
# 方法二：row 行；column 列
# d = wb2.cell(row=4, column=2, value=10)

d = ws[2]
data = [cell.value for cell in d]
print(data)


file = '../excelData.xlsxs'

print(file.endswith('.xlsx'))


a, b = 1, 2
print(a, b)
